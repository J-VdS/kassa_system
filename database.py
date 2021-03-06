# -*- coding: utf-8 -*-
import sqlite3
import pickle
import os
import datetime
import time
#exports
import csv
import openpyxl

#import time
#import func

def update_dict(oud, nieuw):
    for key in nieuw:
        oud[key] = oud.get(key, 0) + nieuw[key]
    return oud


def CloseIO(db_io):
    ''' Deze functie wordt opgeroepen als het scherm gesloten wordt en/of een error optreedt'''
    db_io[1].close() #conn
    db_io[0].close() #cursor

def OpenIO(db):
    conn = sqlite3.connect(db)
    c = conn.cursor()
    return (conn, c)

def InitProduct(db):
    ''' Maakt de producten tabel en zorgt voor db_io '''
    pass

def InitTabels(db_io):
    conn, c = db_io
    #tables_check aanwezig
    #producten tabel
    c.execute("CREATE TABLE IF NOT EXISTS producten(id INTEGER PRIMARY KEY, type TEXT, naam TEXT, prijs INTEGER, active INTEGER)")
    c.execute("CREATE TABLE IF NOT EXISTS totalen(id INTEGER PRIMARY KEY, bestelling BLOB, open INTEGER, prijs INTEGER, naam TEXT, betaalwijze TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS orders(klantid INTEGER, bestelling BLOB, tijd TEXT, unixtijd INTEGER, hash TEXT, types TEXT)")

    #print("--- db loaded ---")

    conn.commit()
   
    
def AddProduct(db_io, type, naam, prijs, active):
    '''
        
        voegt een product toe aan de producten tabel
        return values:
            0: succes
            -1: reeds aanwezig
            -2: error
    '''
    conn, c  = db_io
    c.execute("SELECT naam FROM producten")
    try:
        if naam in [row[0] for row in c.fetchall()]:
            #REEDS AANWEZIG
            return -1
        else:
            c.execute("INSERT INTO producten (naam, type, prijs, active) VALUES(?,?,?,?)",(naam, type, prijs, active))
            conn.commit()
            return 0
    except Exception as e:
        print("\n\n --- AddProduct error ---\n\n" + str(e))
        return -2
    

def getAllProduct(db_io):
    conn, c = db_io
    c.execute("SELECT type, naam, prijs, active FROM producten ORDER BY naam COLLATE NOCASE ASC")
    data = c.fetchall()
    return data


def getAllProductClient(db_io):
    conn, c = db_io
    c.execute("SELECT type, naam FROM producten WHERE active = ?", ('1'),)
    data = c.fetchall()
    return data


def getAllProductKassa(db_io):
    conn, c = db_io
    c.execute("SELECT type, naam, prijs FROM producten WHERE active = 1 OR active = 2 ORDER BY naam COLLATE NOCASE ASC")
    data = c.fetchall()
    return data #lijst bestaande uit tuples


def getTypes(db_io):
    conn, c = db_io
    c.execute("SELECT DISTINCT type FROM producten")
    data = c.fetchall()
    return data


def editProduct(db_io, naam, type, prijs, active):
    conn, c = db_io
    c.execute("SELECT * FROM producten WHERE naam = ?", (naam,))
    if not c.fetchone():
        return -1 #product niet in db
    else:
        c.execute("UPDATE producten SET type = ?, prijs = ?, active = ? WHERE naam = ?", (type, prijs, active, naam))
        conn.commit()
        return 0
    

def deleteProduct(db_io, naam):
    conn, c = db_io
    c.execute("SELECT * FROM producten WHERE naam = ?", (naam,))
    if not c.fetchone():
        return -1 #product niet in db
    else:
        c.execute("DELETE FROM producten WHERE naam = ?", (naam,))
        conn.commit()
        return 0
    
    
def zichtProduct(db_io, naam, active):
    conn, c = db_io
    c.execute("SELECT * FROM producten WHERE naam = ?", (naam,))
    if not c.fetchone():
        return -1
    else:
        c.execute("UPDATE producten SET active = ? WHERE naam = ?", (active, naam))
        conn.commit()
        return 0
    
#bestellingen
def addBestelling(db_io, info, bestelling):
    '''
        db_io (tuple): connectie met db
        info (dict): die info bevat over de klant en de verkoper
        bestelling (dict): nieuwe bestelling
        
        return:
            1  - nieuwe rekening wordt geopend
            0  - bestelling wordt toegevoegd
            -1 - bestelling is al gesloten
    '''
    conn, c = db_io
    c.execute("SELECT bestelling, open FROM totalen WHERE id = ?", (info["id"],))
    data = c.fetchone()
    if not data:
        #maak een nieuw ID aan
        bst = pickle.dumps(bestelling)
        c.execute("INSERT INTO totalen (id, naam, bestelling, open) VALUES (?,?,?,?)", (info["id"], str(info["naam"]), bst, 1))
        conn.commit()
        return 1
    else:
        #bestelling is gesloten
        if data[1] == 0:
            return -1
        bst = pickle.dumps(update_dict(pickle.loads(data[0]), bestelling))
        c.execute("UPDATE totalen SET bestelling = ? WHERE id = ?", (bst, info["id"]))
        conn.commit()
        return 0


def addBestellingID(db_io, ID, naam):
    conn, c = db_io
    c.execute("SELECT bestelling FROM totalen WHERE id = ?", (ID,))
    data = c.fetchone()
    if not data:
        c.execute("INSERT INTO totalen (id, naam, bestelling, open) VALUES (?,?,?,?)", (ID, naam, pickle.dumps({}), 1))
        conn.commit()
        return 1
    else:
        #err ID was al in de DB
        return -1
    

def addOrder(db_io, message, ip_poort="DB", types=None, status="???"):
    """
        <> db_io
        <dict> message
        <str> ip_poort: "ip:poort" van de printer
        <str> types: types in verzonden bericht
    
    """
    conn, c = db_io
    tijd = datetime.datetime.now().strftime("%H:%M:%S")
    ID = message["bestelling"]["info"]["id"]
    if types is None:
        raw_types = ["{:<2}".format(str(i)[:2]) for i in message["bestelling"]["BST"].keys()] #ga naar 1 letter types
        types = "".join(raw_types)
    try:
        c.execute("INSERT INTO orders (klantid, bestelling, tijd, unixtijd, hash, types) VALUES (?,?,?,?,?,?)", (ID, pickle.dumps(message["bestelling"]), tijd, int(time.time()), message['hash'], types))
        conn.commit()
        #["TIJD", "ID", "HASH", "IP:POORT", "TYPES", "STATUS"]
        return [tijd, ID, message['hash'], ip_poort, types, status]
    except Exception as e:
        print(e)
        return -1
    

def getOrder(db_io, _id, _hash):
    conn, c = db_io
    c.execute("SELECT bestelling, types, tijd FROM orders WHERE klantid = ? AND hash = ?", (_id, _hash))
    data = c.fetchone()
    if not data:
        #er was geen bestelling met dit klantid en/of hash
        return -2
    else:
        return [pickle.loads(data[0])] + list(data[1:])
    

def getOrderLastInfo(db_io, _id):
    conn, c = db_io
    c.execute("SELECT tijd, hash, bestelling FROM orders WHERE klantid = ? ORDER BY unixtijd DESC", (_id,))
    data = c.fetchone()
    if not data:
        return ["N/A", "N/A", "N/A"]
    else:
        #bestelling
        tafel = pickle.loads(data[2])['info'].get('tafel', 'N/A')
        return data[:2]+(tafel,)


def getBestelling(db_io, ID):
    conn, c = db_io
    c.execute("SELECT bestelling FROM totalen WHERE id = ?", (ID,))
    data = c.fetchone()
    if not data:
        #er was geen bestelling met dit ID!
        return -1
    else:
        return pickle.loads(data[0]) #{product:aantal}


def getIDs(db_io):
    conn, c = db_io
    c.execute("SELECT id FROM totalen WHERE open = 1 ORDER BY id COLLATE NOCASE ASC")
    data = c.fetchall()
    return data  


def getInfoByID(db_io, ID):
    conn, c = db_io
    c.execute("SELECT naam, open FROM totalen WHERE id = ?", (ID,))
    data = c.fetchone()
    if not data:
        return ("ERROR", "---")
    else:
        return data
    

def delByID(db_io, ID):
    conn, c = db_io
    c.execute("DELETE FROM totalen WHERE id = ?", (ID,))
    conn.commit()


def sluitById(db_io, ID, prijs, bw):
    conn, c = db_io
    c.execute("UPDATE totalen SET open = 0, prijs = ? , betaalwijze = ? WHERE id = ?", (prijs, bw, ID))
    conn.commit()
    

def getTotaalProd(db_io, start=None, end=None, status=0):
    conn, c = db_io
    result = {}
    if start == None and end == None:
        c.execute("SELECT bestelling FROM totalen WHERE open = ?", (status, ))
    elif start and end:
        c.execute("SELECT bestelling FROM totalen WHERE open = ? AND id>=? AND id<=?", (status, start, end))
    elif start:
        c.execute("SELECT bestelling FROM totalen WHERE open = ? AND id>=?", (status, start))
    else:
        c.execute("SELECT bestelling FROM totalen WHERE open = ? AND id<=?", (status, end))
    data = c.fetchall()
    for i in data:
        result = update_dict(result, pickle.loads(i[0])) 
    return result


def getOmzet(db_io, start=None, end=None):
    conn, c = db_io
    result = {"omzet":0}
    if start == None and end == None:
        c.execute("SELECT betaalwijze, SUM(prijs) FROM totalen WHERE open = 0 GROUP BY betaalwijze")
    elif start and end:
        c.execute("SELECT betaalwijze, SUM(prijs) FROM totalen WHERE open = 0 AND id>=? AND id<=? GROUP BY betaalwijze", (start, end))
    elif start:
        c.execute("SELECT betaalwijze, SUM(prijs) FROM totalen WHERE open = 0 AND id>=? GROUP BY betaalwijze", (start,))
    else:
        c.execute("SELECT betaalwijze, SUM(prijs) FROM totalen WHERE open = 0 AND id<=? GROUP BY betaalwijze", (end,))
        
    data = c.fetchall()
    for B, P in data:
        result[B] = P/100
        result["omzet"] += P
    result["omzet"] /= 100
    return result

# export csv
def exportCSV(db_io):#, filename="test.csv"):
    conn, c = db_io
    #check folder
    if not(os.path.isdir("./exports")):
        os.mkdir("./exports")
    PATH = "exports/"+datetime.datetime.now().strftime("%d%m%y@%H-%M-%S_")
    
    c.execute("SELECT type, naam, prijs, active FROM producten ORDER BY naam COLLATE NOCASE ASC")
    
    producten = []
    dict_prod = {}
    bedragen = {}
    
    for T, N, P, A in list(c.fetchall()):
        producten.append([T, N, str(P/100), str(A)])
        dict_prod[N] = 0

    with open(PATH+"productdump.csv", "w", newline='') as f:
        writer = csv.writer(f, delimiter=',')
        #TODO replace . met ,
        writer.writerow(["type", "naam", "prijs", "zichtbaar"])
        for rij in producten:
            writer.writerow(rij)
    del producten
    
    #bestellingen
    c.execute("SELECT ID, naam, open, prijs, betaalwijze, bestelling FROM totalen ORDER BY id ASC")
    data = c.fetchall()
    
    with open(PATH+"bestellingen.csv", "w", newline='') as f:
        writer = csv.writer(f, delimiter=',')
        writer.writerow(["ID", "naam", "open", "prijs", "betaalwijze",  "#", *dict_prod.keys()])
        for ID, N, O, P, B, best in data:
            best = update_dict(dict_prod.copy(), pickle.loads(best))
            #https://stackoverflow.com/questions/35694303/convert-array-of-int-to-array-of-chars-python
            if P:
                writer.writerow([ID, N, str(O), str(P/100), B, "#", *list(map(str, best.values()))])
                bedragen[B] = bedragen.get(B, 0) + P
            else:
                writer.writerow([ID, N, str(O), str(P), B, "#", *list(map(str, best.values()))])
    #ontvangen bedragen
    with open(PATH+"bedragen.csv", "w", newline='') as f:
        writer = csv.writer(f, delimiter=',')
        writer.writerow(["methode", "bedrag (in €)"])
        for key in bedragen:
            writer.writerow([key, str(bedragen[key]/100).replace('.',',')])
            
    #resume
    data = update_dict(dict_prod, getTotaalProd(db_io)) #alle gesloten 
    
    with open(PATH+"bestelde_aantallen.csv", "w", newline='') as f:
        writer = csv.writer(f, delimiter=',')
        writer.writerow(["product", "aantal"])
        for key in data:
            writer.writerow([key, str(data[key])])
            
    
#export xlsx
def exportXLSX(db_io):
    conn, c = db_io
    wb = openpyxl.Workbook()
    
    #productdump
    ws = wb.active
    ws.title = "producten_db"
    
    c.execute("SELECT type, naam, prijs, active FROM producten ORDER BY naam COLLATE NOCASE ASC")
    
    dict_prod = {}
    bedragen = {}
    
    ws.append(("type", "naam", "prijs", "zichtbaar"))
    for T, N, P, A in list(c.fetchall()):
        dict_prod[N] = 0
        ws.append((T, N, P/100, A))
    
    #bestellingen
    ws2 = wb.create_sheet(title="bestellingen")
    
    c.execute("SELECT ID, naam, open, prijs, betaalwijze, bestelling FROM totalen ORDER BY id ASC")
    data = c.fetchall()
    
    ws2.append(("ID", "naam", "open", "prijs", "betaalwijze",  " ", *dict_prod.keys()))
    for ID, N, O, P, B, best in data:
        best = update_dict(dict_prod.copy(), pickle.loads(best))
        if P:
            ws2.append((ID, N, O, P/100, B, " ", *best.values()))
            bedragen[B] = bedragen.get(B, 0) + P
        else:
            ws2.append((ID, N, O, P, B, " ", *best.values())) #P == None
            
    #ontvangen bedragen
    ws3 = wb.create_sheet(title="bedragen")
    
    ws3.append(("methode", "bedrag (in €)"))
    for key in bedragen:
        ws3.append((key, bedragen[key]/100))
        
    #resume
    ws4 = wb.create_sheet(title="verkochte aantallen")
    
    data = update_dict(dict_prod, getTotaalProd(db_io)) #alle gesloten 
    
    ws4.append(("product", "aantal"))
    for key in data:
        ws4.append((key, data[key]))
    
    #save
    if not(os.path.isdir("./exports")):
        os.mkdir("./exports")
    wb.save(filename = "exports/"+datetime.datetime.now().strftime("%d%m%y@%H-%M-%S_")+"export.xlsx")
    
    
def importCSV(file, ret_d):
    try:
        with open(file, newline='') as infile:
            reader = csv.reader(infile, delimiter=',')
            next(reader)
            for row in reader:
                if len(row) != 2:
                    return -1
                K, A = row
                ret_d[K] = A
        return 0
    except Exception as e:
        print(e)
        return -2


def importXLSX(file, ret_d):
    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb["verkochte aantallen"]
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            ret_d[row[0]] = row[1]
        return 0
    except:
        return -1
    